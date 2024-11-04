import csv
import openpyxl
from openpyxl.styles import PatternFill, Alignment

def read_hierarchy_csv(filename):
    hierarchy = []
    max_depth = 0
    with open(filename, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        headers = next(reader, None)  # Skip header if present
        for row in reader:
            path = [level.strip() for level in row if level.strip()]
            hierarchy.append(path)
            if len(path) > max_depth:
                max_depth = len(path)
    return hierarchy, max_depth

def build_tree(paths):
    tree = {}
    for path in paths:
        current_level = tree
        for node in path:
            current_level = current_level.setdefault(node, {})
    return tree

def compare_trees(base_tree, new_tree, depth=0):
    comparison_rows = []
    base_nodes = list(base_tree.keys())
    new_nodes = list(new_tree.keys())
    all_nodes = []

    # Use positions from the base hierarchy
    max_length = max(len(base_nodes), len(new_nodes))
    for i in range(max_length):
        base_node = base_nodes[i] if i < len(base_nodes) else None
        new_node = new_nodes[i] if i < len(new_nodes) else None

        if base_node and new_node and base_node == new_node:
            status = 'match'
        elif base_node and new_node:
            # Nodes are different, include both
            status = 'mismatch'
        elif base_node:
            status = 'only_in_base'
        else:
            status = 'only_in_new'

        comparison_rows.append({
            'depth': depth,
            'base_name': base_node,
            'new_name': new_node,
            'status': status
        })

        # Recurse into subtrees
        base_subtree = base_tree.get(base_node, {}) if base_node else {}
        new_subtree = new_tree.get(new_node, {}) if new_node else {}
        comparison_rows.extend(compare_trees(base_subtree, new_subtree, depth + 1))

    return comparison_rows

def output_to_excel(comparison_rows, max_depth, output_filename='comparison_cco_core_en1.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hierarchy Comparison'

    # Define cell styles
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Matching entities
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Only in base
    blue_fill = PatternFill(start_color='CCF2FF', end_color='CCF2FF', fill_type='solid')    # Only in new

    # Total columns = max_depth * 2 + 2 (for two gap columns)
    total_columns = max_depth * 2 + 2

    # Write header
    headers = ['Base Level {}'.format(i+1) for i in range(max_depth)] + ['', ''] + ['New Level {}'.format(i+1) for i in range(max_depth)]
    ws.append(headers)

    # Set column widths
    narrow_width = 3  # Narrow column width for hierarchy levels
    gap_width = 5     # Wider width for gap columns

    for i in range(1, total_columns + 1):
        column_letter = openpyxl.utils.get_column_letter(i)
        if i == max_depth + 1 or i == max_depth + 2:
            # Gap columns
            ws.column_dimensions[column_letter].width = gap_width
        else:
            ws.column_dimensions[column_letter].width = narrow_width

    # Disable text wrapping for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    # Write rows
    for row_data in comparison_rows:
        depth = row_data['depth']
        base_name = row_data.get('base_name')
        new_name = row_data.get('new_name')
        status = row_data.get('status')

        # Initialize the row with empty cells
        row = [''] * total_columns

        # Base hierarchy columns
        if base_name:
            row[depth] = base_name  # Index from 0
        # New hierarchy columns
        if new_name:
            row[max_depth + 2 + depth] = new_name  # +2 for the gap columns

        # Append the row to the worksheet
        ws.append(row)

        # Apply styles
        excel_row_num = ws.max_row
        # Apply fill to entire hierarchy section
        if base_name or status == 'only_in_base':
            for col in range(1, max_depth + 1):
                cell = ws.cell(row=excel_row_num, column=col)
                if status == 'only_in_base':
                    cell.fill = yellow_fill
                elif status == 'match':
                    cell.fill = green_fill
                elif status == 'mismatch':
                    cell.fill = yellow_fill
                cell.alignment = Alignment(wrap_text=False)

        if new_name or status == 'only_in_new':
            for col in range(max_depth + 3, total_columns + 1):
                cell = ws.cell(row=excel_row_num, column=col)
                if status == 'only_in_new':
                    cell.fill = blue_fill
                elif status == 'match':
                    cell.fill = green_fill
                elif status == 'mismatch':
                    cell.fill = blue_fill
                cell.alignment = Alignment(wrap_text=False)

    wb.save(output_filename)

def main():
    # Replace 'base_hierarchy.csv' and 'new_hierarchy.csv' with your actual file names
    base_hierarchy_csv = 'cco.csv'
    new_hierarchy_csv = 'core.csv'

    # Read the hierarchies from CSV files
    base_paths, base_max_depth = read_hierarchy_csv(base_hierarchy_csv)
    new_paths, new_max_depth = read_hierarchy_csv(new_hierarchy_csv)

    # Build trees from paths
    base_tree = build_tree(base_paths)
    new_tree = build_tree(new_paths)

    # Determine the maximum depth across both hierarchies
    max_depth = max(base_max_depth, new_max_depth)

    # Compare the trees
    comparison_rows = compare_trees(base_tree, new_tree)

    # Output to Excel
    output_to_excel(comparison_rows, max_depth)

if __name__ == '__main__':
    main()
